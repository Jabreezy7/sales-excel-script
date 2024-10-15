import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Function to get item size
def get_item_size(item):
    for option in item.get("selectedOptions", []):
        if option.get("name") == "Size":
            return option.get("value", "N/A")
    return "N/A"

# Function to get custom name/number on jersey
def get_custom_name_and_number(item):
    custom_name = "N/A"
    fixed_name = "N/A"

    # Iterate through all the item options and look for the jersey customization options
    for option in item.get("selectedOptions", []):
        if option.get("name") == "Custom name/ number":
            custom_name = option.get("value", "N/A")
        elif option.get("name") == "Name/ Number":
            fixed_name = option.get("value", "N/A")  

    # If the user inputted a custom name then return the custom name and replace the forward slahes preceding the number for clarity
    # The same logic applies to the fixed name
    if (custom_name != "N/A" and custom_name != "Custom name/ number" and custom_name != "No name/ number" and custom_name != ""):
        return custom_name.replace("/", " ")
    if (fixed_name != "N/A" and fixed_name != "Custom name/ number" and fixed_name != "No name/ number" and fixed_name != ""):
        return fixed_name.replace("/", " ")
    
    # If the user did not input a custom or fixed name then return NONE
    return "NONE"

# Store API details
store_id = "111111..."
api_token = "XXXXXXXXXXX..."

# API endpoint to fetch orders
url = f"https://app.ecwid.com/api/v3/{store_id}/orders"

# Set up the headers for authorization
headers = {
    "Authorization": f"Bearer {api_token}"
}

# Function to fetch the order/sales information
def fetch_sales_data():
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return response.json().get("items", [])
    else:
        print("Failed to retrieve data. Status Code:", response.status_code)
        return []

# Fetch the sales data
sales_data = fetch_sales_data()

# Excel file to save data
excel_file = "SalesData.xlsx"

# Excel Header Row
expected_headers = ["Order #", "Image URL", "Size", "Customization", "Customer Name", "Address", "City", "State", "Country", "Postal Code", "Phone Number"]

# Load existing workbook or create a new one
try:
    workbook = load_workbook(excel_file)
    worksheet = workbook.active

    # Clear the excel sheet
    worksheet.delete_rows(1, worksheet.max_row)

except FileNotFoundError:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sales Data"

# Add the header row as the first row in the excel sheet
worksheet.append(expected_headers)

# Define fill styles and border styles
green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
blue_fill = PatternFill(start_color='CCCCFF', end_color='CCCCFF', fill_type='solid')

# Define the style of the black borders
black_border = Border(left=Side(style='thick', color='000000'), 
                      right=Side(style='thick', color='000000'), 
                      top=Side(style='thick', color='000000'), 
                      bottom=Side(style='thick', color='000000'))

current_border = black_border

# This variable is used to track the last used fill color 
# True = Green
# False = Blue
last_color = True 

# Write the sales data
for order in sales_data:
    order_number = order.get("id")
    customer_name = order.get("shippingPerson", {}).get("name", "N/A")
    customer_address = order.get("shippingPerson", {}).get("street", "N/A")
    city = order.get("shippingPerson", {}).get("city", "N/A")
    state = order.get("shippingPerson", {}).get("stateOrProvinceName", "N/A")
    country = order.get("shippingPerson", {}).get("countryName", "N/A")
    postal_code = order.get("shippingPerson", {}).get("postalCode", "N/A")
    phone_number = order.get("shippingPerson", {}).get("phone", "N/A")

    # Define sizes and images arrays to store customer product shirt sizes
    # and image sizes as customers can order more than one item and excel
    # only allows one picture per cell.
    sizes = []
    images = []
    

    start_row = worksheet.max_row + 1

    # Iterate through all of the items
    for item in order.get("items", []):
        images.append(item.get("imageUrl", "No small thumbnail available"))
        sizes.append(get_item_size(item))
        customization = get_custom_name_and_number(item)

        # Add a new row for each item
        row = [
            order_number if len(sizes) == 1 else "",  # Show order number only for the first item
            f'=IMAGE("{images[-1]}")',  # Add image URL to the current row and apply excel =IMAGE() function to turn url to image
            sizes[-1],
            customization,
            customer_name,
            customer_address,
            city,
            state,
            country,
            postal_code,
            phone_number
        ]
        
        # Append the new row containing the item
        worksheet.append(row)

    # Fill in cells with active color 
    current_fill = green_fill if last_color else blue_fill

    end_row = worksheet.max_row  
    for row in range(start_row, end_row + 1):
        for col in range(1, len(expected_headers) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = current_fill
            cell.border = current_border

    # Alternate the color
    last_color = not last_color

# Apply text wrapping and top vertical alignment for all rows
for col in range(1, len(expected_headers) + 1):
    for row in range(2, worksheet.max_row + 1):  # Start from row 2 to skip header
        cell = worksheet.cell(row=row, column=col)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.font = Font(size=12)  # Set font size for content cells

# This can be used to dynamically adjust column widths based on content
# however as we are manually setting the columns to have a certain width this is not needed
# max_column_width = 30
# for col in range(1, len(expected_headers) + 1):
#     max_length = 0
#     for cell in worksheet.iter_rows(min_row=2, min_col=col, max_col=col):
#         for c in cell:
#             if c.value:
#                 max_length = max(max_length, len(str(c.value)))
#     adjusted_width = min(max_length + 2, max_column_width)
#     worksheet.column_dimensions[worksheet.cell(row=1, column=col).column_letter].width = adjusted_width

# Set all columns to have a certain width
fixed_column_width = 30
for col in range(1, len(expected_headers) + 1):
    column_letter = worksheet.cell(row=1, column=col).column_letter
    worksheet.column_dimensions[column_letter].width = fixed_column_width

# Adjust row height dynamically based on content
for row in range(2, worksheet.max_row + 1):
    total_height = 15
    for col in range(1, len(expected_headers) + 1):
        cell_value = worksheet.cell(row=row, column=col).value
        if cell_value:
            num_lines = str(cell_value).count('\n') + 1
            total_height += num_lines * 15
    worksheet.row_dimensions[row].height = total_height


# Set all cells to have font-size 20
font_size = 20
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
    for cell in row:
        cell.font = Font(size=font_size)

# Save the workbook
workbook.save(excel_file)
print(f"Sales data saved to {excel_file}.")
